import streamlit as st
import os
from pytube import YouTube
import yt_dlp
from pydub import AudioSegment
from moviepy import VideoFileClip
from google.cloud import speech
from google.cloud import storage
from concurrent.futures import ThreadPoolExecutor
import base64
import json
import openai
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
from urllib.parse import urlparse, parse_qs
import tiktoken
from PIL import Image

# Timer
if 't' not in st.session_state:
    st.session_state.t = 2  # Default value

# Function to identify YouTube URLs
def is_youtube_url(url):
    try:
        parsed_url = urlparse(url)
        domain = parsed_url.netloc.lower()
        return (
            domain.endswith("youtube.com") or
            domain == "youtu.be"
        )
    except Exception as e:
        st.error(f"Error parsing URL: {e}")
        return False

# Function to identify X.com (Twitter) URLs
def is_xcom_url(url):
    try:
        parsed_url = urlparse(url)
        domain = parsed_url.netloc.lower()
        return (
            domain.endswith("x.com") or
            domain.endswith("twitter.com")
        )
    except Exception as e:
        st.error(f"Error parsing URL: {e}")
        return False

# Function to fetch YouTube video details using yt_dlp
def fetch_youtube_details(video_url):
    try:
        ydl_opts = {
            'quiet': True,
            'skip_download': True,
            'forcejson': True,
        }
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
        
        # Extract relevant details
        video_title = info.get('title', 'No Title')
        thumbnail_url = info.get('thumbnail', '')
        duration = info.get('duration', 0)  # Duration in seconds
        return video_title, thumbnail_url, duration
    except Exception as e:
        st.error(f"Error fetching YouTube video details: {e}")
        return None, None, None

# Function to fetch X.com (Twitter) video details using yt_dlp
def fetch_x_details(video_url):
    try:
        ydl_opts = {
            'quiet': True,
            'skip_download': True,
            'forcejson': True,
        }
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
        
        # Extract relevant details
        video_title = info.get('title', 'No Title')
        thumbnail_url = info.get('thumbnail', '')
        duration = info.get('duration', 0)  # Duration in seconds
        return video_title, thumbnail_url, duration
    except Exception as e:
        st.error(f"Error fetching X.com video details: {e}")
        return None, None, None

# Function to fetch video details based on URL type
def fetch_video_details(video_url):
    if is_youtube_url(video_url):
        return fetch_youtube_details(video_url)
    elif is_xcom_url(video_url):
        return fetch_x_details(video_url)
    else:
        st.error("Unsupported URL. Please provide a YouTube or X.com URL.")
        return None, None, None

# Access the credentials from Streamlit secrets
google_cloud_credentials = st.secrets["google_cloud"]
google_cloud_credentials_dict = dict(google_cloud_credentials)

# Write the credentials to a temporary JSON file
with open("gcloud_temp_credentials.json", "w") as f:
    json.dump(google_cloud_credentials_dict, f)

# Set the environment variable to point to the temporary JSON file
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "gcloud_temp_credentials.json"

# Global variables to store the final transcribed output
transcripted_output = ""

# Function to clean up old video and audio files if they exist
def cleanup_old_files():
    video_path = 'video.mp4'
    audio_path = 'audio_compressed.wav'
    
    # Check if old video file exists and remove it
    if os.path.exists(video_path):
        os.remove(video_path)
        print(f"Removed old video file: {video_path}")
    
    # Check if old audio file exists and remove it
    if os.path.exists(audio_path):
        os.remove(audio_path)
        print(f"Removed old audio file: {audio_path}")

# Function to download YouTube video using yt_dlp
def download_youtube_video(video_url, output_video_path='video.mp4'):
    ydl_opts = {
        'format': 'best',
        'outtmpl': output_video_path,
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        ydl.download([video_url])
    
    print(f"Video downloaded and saved as {output_video_path}")
    return output_video_path

# Function to extract audio from video using moviepy + pydub
def extract_audio_from_video(video_path, output_audio_path='audio_compressed.wav', sample_rate=16000, bitrate='32k'):
    video = VideoFileClip(video_path)
    audio_path_temp = "temp_audio.wav"
    
    # Extract audio using moviepy
    audio = video.audio
    audio.write_audiofile(audio_path_temp, codec='pcm_s16le', fps=44100)  # 44100 Hz
    
    # Convert audio to mono and compress using pydub
    sound = AudioSegment.from_file(audio_path_temp)
    sound = sound.set_channels(1)  # Convert to mono
    sound = sound.set_frame_rate(sample_rate)  # Set sample rate to 16000 Hz
    sound.export(output_audio_path, format="wav", bitrate=bitrate)
    
    print(f"Compressed audio extracted and saved as {output_audio_path}.")
    # Clean up temporary audio file
    os.remove(audio_path_temp)
    
    return output_audio_path

# Function to split the audio into chunks
def split_audio_to_chunks(audio_path, chunk_duration_ms=60000):
    sound = AudioSegment.from_wav(audio_path)
    audio_chunks = []
    for i in range(0, len(sound), chunk_duration_ms):
        chunk = sound[i:i + chunk_duration_ms]
        chunk_path = f"chunk_{i // chunk_duration_ms}.wav"
        chunk.export(chunk_path, format="wav")
        audio_chunks.append(chunk_path)
    
    print(f"Split audio into {len(audio_chunks)} chunks, each of {chunk_duration_ms} ms.")
    return audio_chunks

# Function to upload audio to Google Cloud Storage
def upload_to_gcs(bucket_name, source_file_name, destination_blob_name):
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_filename(source_file_name)
    print(f"File {source_file_name} uploaded to {destination_blob_name}.")

# Function to transcribe audio from Google Cloud Storage
def transcribe_audio_gcs(gcs_uri, language_code='en-US'):
    client = speech.SpeechClient()
    audio = speech.RecognitionAudio(uri=gcs_uri)
    config = speech.RecognitionConfig(
        encoding=speech.RecognitionConfig.AudioEncoding.LINEAR16,
        sample_rate_hertz=16000,
        language_code=language_code,
        enable_automatic_punctuation=True,
    )
    operation = client.long_running_recognize(config=config, audio=audio)
    response = operation.result(timeout=1500)

    # Print and return the transcription
    full_transcript = ""
    for result in response.results:
        full_transcript += f"{result.alternatives[0].transcript}\n"
    
    print(f"Transcription complete for language {language_code}.")
    return full_transcript

# Function to transcribe audio chunks in parallel (Hindi + English)
def transcribe_audio_chunks_in_parallel(bucket_name, chunk_paths, language_code='en-US'):
    with ThreadPoolExecutor() as executor:
        futures = []
        for chunk_path in chunk_paths:
            destination_blob_name = os.path.basename(chunk_path)
            upload_to_gcs(bucket_name, chunk_path, destination_blob_name)
            gcs_uri = f"gs://{bucket_name}/{destination_blob_name}"
            futures.append(executor.submit(transcribe_audio_gcs, gcs_uri, language_code))
        transcriptions = [future.result() for future in futures]
    return " ".join(transcriptions)

# If we take a local video file, transcribe here
def transcribe_local_video(local_video_path):
    global transcripted_output
    cleanup_old_files()
    
    # Extract and split audio into chunks
    audio_path = extract_audio_from_video(local_video_path, sample_rate=16000, bitrate='32k')
    audio_chunks = split_audio_to_chunks(audio_path, chunk_duration_ms=60000)
    
    bucket_name = 'hackathon_police'  # your GCS bucket
    transcript_hindi = transcribe_audio_chunks_in_parallel(bucket_name, audio_chunks, language_code='hi-IN')
    st.session_state.t = 1
    transcript_english = transcribe_audio_chunks_in_parallel(bucket_name, audio_chunks, language_code='en-US')
    
    transcripted_output = f"Hindi Transcription:\n{transcript_hindi}\n\nEnglish Transcription:\n{transcript_english}"
    
    for chunk_path in audio_chunks:
        os.remove(chunk_path)

# If we take a local audio file, transcribe here (skips extracting audio from video)
def transcribe_local_audio(local_audio_path):
    global transcripted_output
    cleanup_old_files()
    
    # Directly split the provided audio
    audio_chunks = split_audio_to_chunks(local_audio_path, chunk_duration_ms=60000)
    
    bucket_name = 'hackathon_police'
    transcript_hindi = transcribe_audio_chunks_in_parallel(bucket_name, audio_chunks, language_code='hi-IN')
    st.session_state.t = 1
    transcript_english = transcribe_audio_chunks_in_parallel(bucket_name, audio_chunks, language_code='en-US')
    
    transcripted_output = f"Hindi Transcription:\n{transcript_hindi}\n\nEnglish Transcription:\n{transcript_english}"
    
    for chunk_path in audio_chunks:
        os.remove(chunk_path)

# Handle the entire workflow for youtube/x.com
def transcribe_youtube_video(video_url):
    global transcripted_output
    cleanup_old_files()
    
    # 1. Download
    video_path = download_youtube_video(video_url)
    
    # 2. Extract and split audio
    audio_path = extract_audio_from_video(video_path, sample_rate=16000, bitrate='32k')
    audio_chunks = split_audio_to_chunks(audio_path, chunk_duration_ms=60000)
    
    # 3. Transcribe
    bucket_name = 'hackathon_police'
    transcript_hindi = transcribe_audio_chunks_in_parallel(bucket_name, audio_chunks, language_code='hi-IN')
    st.session_state.t = 1
    transcript_english = transcribe_audio_chunks_in_parallel(bucket_name, audio_chunks, language_code='en-US')
    
    transcripted_output = f"Hindi Transcription:\n{transcript_hindi}\n\nEnglish Transcription:\n{transcript_english}"
    
    for chunk_path in audio_chunks:
        os.remove(chunk_path)

# Classification function
def classify_content(rp_percentage, rc_percentage):
    if rp_percentage >= 70 or rc_percentage >= 70:
        return "red"
    elif 40 <= rp_percentage < 70 or 40 <= rc_percentage < 70:
        return "yellow"
    else:
        return "green"

# Helper to count tokens (for dynamic model selection)
def count_tokens(text: str, model: str = "gpt-4") -> int:
    encoding = tiktoken.encoding_for_model(model)
    tokens = encoding.encode(text)
    return len(tokens)

# GPT Analysis
def get_analysis_with_api_key(transcript):
    openai.api_key = st.secrets["default"]["OPENAI_API_KEY"]
    
    prompt = f"""
    You are tasked with analyzing transcripts of speeches or text content that might include both Hindi and English sections. The transcript has been processed using two separate speech-to-text APIs: one for Hindi and one for English. Analyze the provided transcript carefully, understanding both languages, and return the analysis based on the following five parameters. The transcript might contain mixed Hindi and English parts, so ensure you identify the language for each section and analyze the radical or religiously inflammatory language accordingly.

    ### Parameters to analyze:
    1. *Lexical Analysis*: Identify radical or religious terminology in both Hindi and English, including exclusionary language (e.g., "us vs. them"), calls to action, or divisive rhetoric.
    2. *Emotion and Sentiment in Speech*: Analyze the tone and sentiment in both languages. Look for negative emotions like anger or fear, which may incite followers or condemn opposing groups.
    3. *Speech Patterns and Intensity*: Identify the use of high volume, repetition, or urgency in either language to emphasize points, typical in radical speech.
    4. *Use of Religious Rhetoric*: Look for quotes from religious texts, apocalyptic themes, or divine rewards/punishments, considering the context in both Hindi and English.
    5. *Frequency of Commands and Directives*: Examine the frequency of explicit calls to action (physical or ideological), in both languages.

    ### Standard Output Format:
    Return the analysis in this structured format:

    <u><b>Final Assessment</b></u>: 
    <br> 
    <b><span style="color:red">Radical Probability</span></b>: [Insert percentage here];  
    <b><span style="color:red">Radical Content</span></b>: [Insert percentage here].

    [Separator]

    <b>Lexical Analysis</b>:  
    [Insert analysis here]

    <b>Emotion and Sentiment in Speech</b>:  
    [Insert analysis here]

    <b>Speech Patterns and Intensity</b>:  
    [Insert analysis here]

    <b>Use of Religious Rhetoric</b>:  
    [Insert analysis here]

    <b>Frequency of Commands and Directives</b>:  
    [Insert analysis here]

    Transcript: {transcript}
    """

    token_count = count_tokens(prompt, model="gpt-4")
    if token_count > 8000:
        chosen_model = "gpt-3.5-turbo"
    else:
        chosen_model = "gpt-4"

    response = openai.ChatCompletion.create(
        model=chosen_model,
        messages=[
            {"role": "system", "content": "You are an assistant that analyzes transcripts for radical content."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=1000
    )
    result = response['choices'][0]['message']['content']

    # Split the response into two parts based on "[Separator]" (or fallback)
    try:
        if "[Separator]" in result:
            final_assessment, analysis = result.split("[Separator]", 1)
        elif "---" in result:
            final_assessment, analysis = result.split("---", 1)
        else:
            st.error("Error analyzing: separator not found in the output.")
            final_assessment, analysis = "", ""
    except ValueError:
        st.error("Error analyzing: Unexpected response from model.")
        final_assessment, analysis = "", ""
    
    return final_assessment, analysis

# Extract radical probabilities
def extract_percentages(analysis_text):
    lines = analysis_text.splitlines()

    # Search for lines containing the relevant phrases
    rp_line = [line for line in lines if "Radical Probability" in line]
    rc_line = [line for line in lines if "Radical Content" in line]

    # Helper
    def convert_to_percentage(text):
        text = text.lower()
        if "low" in text:
            return 20
        elif "medium" in text:
            return 50
        elif "high" in text:
            return 80
        else:
            try:
                return int(text.replace("%", "").strip())
            except ValueError:
                return 0

    rp_percentage = convert_to_percentage(rp_line[0].split(":")[1].strip()) if rp_line else 0
    rc_percentage = convert_to_percentage(rc_line[0].split(":")[1].strip()) if rc_line else 0
    return rp_percentage, rc_percentage

# Extract relevant analysis parts (optional if you want them individually)
def extract_analysis_parts(analysis_text):
    lines = analysis_text.splitlines()
    
    def extract_section(section_name):
        try:
            return next(line for line in lines if section_name in line).split(":")[1].strip()
        except StopIteration:
            return "Not available"

    lexical = extract_section("*Lexical Analysis*")
    emotion = extract_section("*Emotion and Sentiment in Speech*")
    speech_patterns = extract_section("*Speech Patterns and Intensity*")
    religious_rhetoric = extract_section("*Use of Religious Rhetoric*")
    commands = extract_section("*Frequency of Commands and Directives*")
    
    return lexical, emotion, speech_patterns, religious_rhetoric, commands

# Append to Excel (optional usage if needed)
def append_to_csv(transcript, analysis_text, rp_percentage, rc_percentage):
    file_path = "analysis_results.xlsx"
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Transcript",
            "Lexical Analysis",
            "Emotion and Sentiment",
            "Speech Patterns",
            "Religious Rhetoric",
            "Commands",
            "Radical Probability",
            "Radical Content",
            "Classification"
        ])
        workbook.save(file_path)

    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    classification = classify_content(rp_percentage, rc_percentage)
    lexical, emotion, speech_patterns, religious_rhetoric, commands = extract_analysis_parts(analysis_text)
    
    row_data = [
        transcript,
        lexical,
        emotion,
        speech_patterns,
        religious_rhetoric,
        commands,
        rp_percentage,
        rc_percentage,
        classification
    ]
    sheet.append(row_data)

    fill_color = {
        "red": "FF0000",
        "yellow": "FFFF00",
        "green": "00FF00"
    }
    fill = PatternFill(
        start_color=fill_color[classification],
        end_color=fill_color[classification],
        fill_type="solid"
    )
    for cell in sheet[sheet.max_row]:
        cell.fill = fill
    
    workbook.save(file_path)

# Create downloads dir if not exist
directory = 'downloads/'
if not os.path.exists(directory):
    os.makedirs(directory)

# Streamlit page config
st.set_page_config(
    page_title="Cyber Crime Department - Goa Police",
    page_icon="https://upload.wikimedia.org/wikipedia/en/d/dd/Emblem_of_Goa_Police.png",
    layout="wide"
)

# Custom CSS
custom_style = """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    body {
        background-color: white !important;
    }
    .navbar {
        overflow: hidden;
        background-color: #D8D4B8;
        position: fixed;
        top: 0;
        left: 0;  
        right: 0; 
        width: 100%;
        z-index: 1000;
    }
    .navbar img {
        float: left;
        display: inline-block;
        padding: 14px 10px;
        height: 128px;
    }
    .navbar .title {
        color: white;
        float: left;
        font-size: 64px;
        margin: 0;
        padding: 25px 10px;
        font-family: 'Times New Roman', serif;
        font-weight: bold;
        text-transform: uppercase;
        letter-spacing: 1.5px;
    }
    .subtitle {
        color: #000000;
        text-align: center;
        font-size: 48px;
        font-family: 'Cambria', serif;
        font-weight: bold;
        padding: 20px 0;
        padding-top: 60px;
        text-transform: capitalize;
    }
    .report {
        color: #000000; 
        text-align: justify;
        font-size: 18px;
    }
    .assess {
        color: #000000; 
        text-align: justify;
        font-size: 18px;
    }
    @media screen and (max-width: 600px) {
        .navbar .title {
            font-size: 18px;
        }
        .subtitle {
            font-size: 24px;
        }
    }
    </style>
"""
st.markdown(custom_style, unsafe_allow_html=True)

# Navbar
st.markdown("""
    <div class="navbar">
        <img src="https://upload.wikimedia.org/wikipedia/en/d/dd/Emblem_of_Goa_Police.png" alt="Goa Police Logo">
        <div class="title">Goa Police</div>
    </div>
""", unsafe_allow_html=True)

# Page title
st.markdown('<div class="subtitle">Radical and Religious Content Analyzer</div>', unsafe_allow_html=True)
st.markdown("<br><br><br>", unsafe_allow_html=True)

# -- UI inputs --
url = st.text_input(
    "Paste YouTube/X.com URL here",
    placeholder='https://www.youtube.com/ or https://x.com/'
)

st.markdown('<div style="text-align: center;"> <p>OR</p></div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Upload a local file (Video or Audio)",
    type=["mp4", "mov", "avi", "m4a", "mp3", "wav"]  # add more audio types as needed
)

st.markdown('<div style="text-align: center;"> <p>OR</p></div>', unsafe_allow_html=True)

# Direct Transcript Text
transcript_text = st.text_area("Enter or paste transcript text (optional). If provided, analysis will skip any audio/video steps.")

analyze_button = st.button("Analyze")

# -- Logic to handle user actions --
if analyze_button:

    # 1) If user provided transcript text directly, skip all audio/video steps
    if transcript_text.strip():
        transcripted_output = transcript_text.strip()
        with st.spinner("Analyzing transcript text..."):
            final_assess, analysis = get_analysis_with_api_key(transcripted_output)
        
        # Display results
        final_assess_html = final_assess.replace('\n\n', '<br><br>').replace('\n', '<br>')
        analysis_html = analysis.replace('\n\n', '<br><br>').replace('\n', '<br>')
        st.markdown(f"<div class='assess' style='white-space: pre-wrap;'>{final_assess_html}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='report' style='white-space: pre-wrap;'>{analysis_html}</div>", unsafe_allow_html=True)
        
        rp_percentage, rc_percentage = extract_percentages(analysis)
        # append_to_csv(transcripted_output, analysis, rp_percentage, rc_percentage)  # optional
        # st.dataframe(pd.read_excel('analysis_results.xlsx').tail())

    # 2) Otherwise, check if a URL was provided
    elif url:
        with st.spinner(f"Fetching video details... Estimated time: {st.session_state.t} mins"):
            video_title, thumbnail_url, duration = fetch_video_details(url)
            if video_title:
                # Convert duration to mm:ss
                minutes, seconds = divmod(duration, 60)
                video_length = f"{int(minutes)}:{int(seconds):02d}"
                
                col1, col2 = st.columns([1, 1.5], gap="small")
                with col1:
                    st.image(thumbnail_url, caption="Video Thumbnail", use_container_width=True)
                with col2:
                    st.markdown('<h2 style="font-weight:bold;">Video Details</h2>', unsafe_allow_html=True)
                    st.write(f"**Title:** {video_title}")
                    st.write(f"**Length:** {video_length}")
                
                # Transcribe
                transcribe_youtube_video(url)
                # Analyze
                transcript = transcripted_output
                final_assess, analysis = get_analysis_with_api_key(transcript)
                
                final_assess_html = final_assess.replace('\n\n', '<br><br>').replace('\n', '<br>')
                analysis_html = analysis.replace('\n\n', '<br><br>').replace('\n', '<br>')
                
                with col2:
                    st.markdown(f"<div class='assess' style='white-space: pre-wrap;'>{final_assess_html}</div>", unsafe_allow_html=True)
                st.markdown(f"<div class='report' style='white-space: pre-wrap;'>{analysis_html}</div>", unsafe_allow_html=True)
                
                rp_percentage, rc_percentage = extract_percentages(analysis)
                # append_to_csv(transcript, analysis, rp_percentage, rc_percentage)
                # st.dataframe(pd.read_excel('analysis_results.xlsx').tail())

            else:
                st.error("Unable to fetch video details. Please check the URL.")

    # 3) Otherwise, if a local file is provided
    elif uploaded_file:
        col1, col2 = st.columns([1, 1.5], gap="small")
        file_extension = uploaded_file.name.split('.')[-1].lower()

        # Distinguish between video vs audio
        video_formats = ["mp4", "mov", "avi"]
        audio_formats = ["m4a", "mp3", "wav"]

        temp_path = f"uploaded_temp.{file_extension}"
        with open(temp_path, "wb") as f:
            f.write(uploaded_file.read())

        if file_extension in video_formats:
            # Handle local video
            with st.spinner(f"Processing local video file... Estimated time: {st.session_state.t} mins"):
                # Get some details
                try:
                    clip = VideoFileClip(temp_path)
                    duration = int(clip.duration)
                    frame = clip.get_frame(0)
                    image_pil = Image.fromarray(frame)
                    thumb_path = "local_thumbnail.png"
                    image_pil.save(thumb_path)
                    clip.close()
                except Exception as e:
                    st.error(f"Error processing video file: {e}")
                    duration = 0
                    thumb_path = None
                
                minutes, seconds = divmod(duration, 60)
                video_length = f"{int(minutes)}:{int(seconds):02d}"
                
                with col1:
                    if thumb_path and os.path.exists(thumb_path):
                        st.image(thumb_path, caption="Video Thumbnail", use_container_width=True)
                    else:
                        st.image("https://via.placeholder.com/150", caption="No Thumbnail", use_container_width=True)
                with col2:
                    st.markdown('<h2 style="font-weight:bold;">Video Details</h2>', unsafe_allow_html=True)
                    st.write(f"**Title:** {uploaded_file.name}")
                    st.write(f"**Length:** {video_length}")

                # Transcribe
                transcribe_local_video(temp_path)
                # Analyze
                final_assess, analysis = get_analysis_with_api_key(transcripted_output)
                
                final_assess_html = final_assess.replace('\n\n', '<br><br>').replace('\n', '<br>')
                analysis_html = analysis.replace('\n\n', '<br><br>').replace('\n', '<br>')
                
                with col2:
                    st.markdown(f"<div class='assess' style='white-space: pre-wrap;'>{final_assess_html}</div>", unsafe_allow_html=True)
                st.markdown(f"<div class='report' style='white-space: pre-wrap;'>{analysis_html}</div>", unsafe_allow_html=True)
                
                rp_percentage, rc_percentage = extract_percentages(analysis)
                # append_to_csv(transcripted_output, analysis, rp_percentage, rc_percentage)
                
                os.remove(temp_path)
                if thumb_path and os.path.exists(thumb_path):
                    os.remove(thumb_path)

        elif file_extension in audio_formats:
            # Handle local audio: skip video extraction
            with st.spinner(f"Processing local audio file... Estimated time: {st.session_state.t} mins"):
                # There's no built-in "thumbnail" concept for audio
                with col1:
                    st.image("https://via.placeholder.com/150", caption="Audio File", use_container_width=True)
                with col2:
                    st.markdown('<h2 style="font-weight:bold;">Audio Details</h2>', unsafe_allow_html=True)
                    st.write(f"**Title:** {uploaded_file.name}")

                # Transcribe
                transcribe_local_audio(temp_path)
                # Analyze
                final_assess, analysis = get_analysis_with_api_key(transcripted_output)
                
                final_assess_html = final_assess.replace('\n\n', '<br><br>').replace('\n', '<br>')
                analysis_html = analysis.replace('\n\n', '<br><br>').replace('\n', '<br>')
                
                with col2:
                    st.markdown(f"<div class='assess' style='white-space: pre-wrap;'>{final_assess_html}</div>", unsafe_allow_html=True)
                st.markdown(f"<div class='report' style='white-space: pre-wrap;'>{analysis_html}</div>", unsafe_allow_html=True)
                
                rp_percentage, rc_percentage = extract_percentages(analysis)
                # append_to_csv(transcripted_output, analysis, rp_percentage, rc_percentage)

                os.remove(temp_path)

        else:
            st.error("Unsupported file format. Please provide a valid video (mp4/mov/avi) or audio file (m4a/mp3/wav).")

    else:
        st.error("No input provided. Please enter a URL, upload a file, or paste transcript text.")

else:
    st.write("Enter a valid YouTube or X.com URL, or upload a valid file, or paste transcript text. Then click 'Analyze'.")
